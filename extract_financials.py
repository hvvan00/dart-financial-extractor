"""Extract exactly three clean financial statements from a DART filing."""

from __future__ import annotations

import argparse
import html
import math
import numbers
import os
import re
import sys
import tempfile
import time
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dart_link import InvalidDisclosureReference, parse_disclosure_reference
from pdf_financials import (
    DART_MAIN_URL,
    PdfExtractionError,
    download_filing_pdf,
    extract_pdf_statements,
    resolve_document_number,
)


STATEMENT_SPECS = (
    ("재무상태표", "get_financial_statement"),
    ("손익계산서", "get_income_statement"),
    ("현금흐름표", "get_cash_flows"),
)
STATEMENT_NAMES = tuple(name for name, _ in STATEMENT_SPECS)
SCOPE_LABELS = {
    "consolidated": "연결",
    "separate": "별도",
}
REPORT_TYPES = ("사업보고서", "반기보고서", "분기보고서", "감사보고서")
_INVALID_FILENAME_CHARACTERS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_HTML_TAG_PATTERN = re.compile(r"<[^>]+>")
_DART_CONNECTION_ATTEMPTS = 3
_DART_RETRY_DELAYS = (5, 15)


class FinancialExtractionError(RuntimeError):
    """Base error for expected extraction failures."""


class MissingApiKeyError(FinancialExtractionError):
    """Raised when DART_API_KEY is absent."""


class MissingXbrlError(FinancialExtractionError):
    """Raised when the filing does not provide a usable XBRL document."""


class MissingStatementError(FinancialExtractionError):
    """Raised when one or more required statements cannot be extracted."""


@dataclass(frozen=True)
class FilingMetadata:
    """The disclosure metadata used only to create a readable output filename."""

    company_name: str
    report_type: str
    year_month: str

    @property
    def filename_stem(self) -> str:
        company = _safe_filename_part(self.company_name)
        return f"{company}_{self.report_type}_{self.year_month}"


@dataclass(frozen=True)
class FilingExtraction:
    """The statements and provenance extracted from one DART filing."""

    metadata: FilingMetadata
    statements: Mapping[str, pd.DataFrame]
    selected_scope: str
    receipt_number: str
    source: str


def _safe_filename_part(value: str) -> str:
    cleaned = _INVALID_FILENAME_CHARACTERS.sub("_", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    cleaned = re.sub(r"_+", "_", cleaned)
    if not cleaned:
        raise FinancialExtractionError("회사명을 안전한 파일명으로 만들 수 없습니다.")
    return cleaned


def parse_filing_metadata_html(html_text: str) -> FilingMetadata:
    """Read company, report type, and filing year/month from a DART page title."""

    title_match = re.search(
        r"<title[^>]*>(?P<title>.*?)</title>",
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    title = title_match.group("title") if title_match else html_text
    title = html.unescape(_HTML_TAG_PATTERN.sub("", title))
    title = " ".join(title.replace("\u00a0", " ").split())

    report_pattern = "|".join(REPORT_TYPES)
    date_pattern = (
        r"(?P<year>(?:19|20)\d{2})[.\-/년]\s*"
        r"(?P<month>\d{1,2})"
        r"(?:[.\-/월]\s*(?:\d{1,2}일?)?)?"
    )
    patterns = (
        re.compile(
            rf"(?P<company>.+?)\s*/\s*"
            rf"(?P<report>{report_pattern})\s*/\s*{date_pattern}"
        ),
        re.compile(
            rf"\[(?P<company>[^\]]+)\]\s*"
            rf"(?P<report>{report_pattern})\s*\(\s*{date_pattern}\s*\)"
        ),
    )
    for pattern in patterns:
        match = pattern.search(title)
        if not match:
            continue
        company_name = match.group("company").strip()
        report_type = match.group("report")
        month = int(match.group("month"))
        if not 1 <= month <= 12:
            break
        return FilingMetadata(
            company_name=company_name,
            report_type=report_type,
            year_month=f"{match.group('year')}.{month:02d}",
        )

    raise FinancialExtractionError(
        "공시 페이지에서 회사명, 보고서 종류, 연월을 찾을 수 없습니다."
    )


def resolve_filing_metadata(
    receipt_number: str,
    dart_module: Any,
) -> FilingMetadata:
    """Download the DART disclosure page and resolve output filename metadata."""

    try:
        response = dart_module.utils.request.get(
            url=DART_MAIN_URL,
            payload={"rcpNo": receipt_number},
            timeout=120,
        )
        response.raise_for_status()
    except Exception as exc:
        raise FinancialExtractionError(
            f"공시 파일명 정보를 확인하지 못했습니다: {exc}"
        ) from exc
    return parse_filing_metadata_html(response.text)


def require_api_key(environ: Mapping[str, str] | None = None) -> str:
    """Read and validate the API key from DART_API_KEY only."""

    source = os.environ if environ is None else environ
    api_key = source.get("DART_API_KEY", "").strip()
    if not api_key:
        raise MissingApiKeyError(
            "DART_API_KEY가 없습니다. GitHub 저장소의 Actions secret에 "
            "DART_API_KEY를 등록하세요."
        )
    return api_key


def _import_dart_fss() -> Any:
    try:
        import dart_fss as dart
    except ImportError as exc:  # pragma: no cover - dependency exists in Actions
        raise FinancialExtractionError(
            "dart-fss가 설치되지 않았습니다. requirements.txt 의존성을 설치하세요."
        ) from exc
    return dart


def _exception_chain(exc: BaseException) -> list[BaseException]:
    chain: list[BaseException] = []
    current: BaseException | None = exc
    seen: set[int] = set()
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        chain.append(current)
        current = current.__cause__ or current.__context__
    return chain


def _is_transient_network_error(exc: BaseException) -> bool:
    """Return whether an exception represents a retryable connection failure."""

    retryable_names = {
        "ConnectTimeout",
        "ConnectTimeoutError",
        "ConnectionError",
        "ConnectionResetError",
        "MaxRetryError",
        "NewConnectionError",
        "ProtocolError",
        "ReadTimeout",
        "ReadTimeoutError",
        "Timeout",
        "TimeoutError",
    }
    retryable_messages = (
        "timed out",
        "timeout",
        "max retries exceeded",
        "temporary failure in name resolution",
        "name or service not known",
        "connection reset",
        "connection aborted",
        "connection refused",
        "network is unreachable",
        "remote disconnected",
    )
    for error in _exception_chain(exc):
        if error.__class__.__name__ in retryable_names:
            return True
        message = str(error).casefold()
        if any(marker in message for marker in retryable_messages):
            return True
    return False


def _set_dart_api_key_with_retry(dart: Any, api_key: str) -> None:
    """Configure dart-fss, retrying only transient Open DART network errors."""

    for attempt in range(1, _DART_CONNECTION_ATTEMPTS + 1):
        try:
            dart.set_api_key(api_key=api_key)
            return
        except Exception as exc:
            if not _is_transient_network_error(exc):
                raise FinancialExtractionError(
                    "DART_API_KEY 인증에 실패했습니다. GitHub 저장소의 "
                    "Actions secret 값이 올바른지 확인하세요."
                ) from exc
            if attempt == _DART_CONNECTION_ATTEMPTS:
                raise FinancialExtractionError(
                    "Open DART 서버 연결이 계속 지연되어 3번 시도했지만 "
                    "인증하지 못했습니다. 입력이나 API 키 문제가 아닐 수 있으므로 "
                    "잠시 후 GitHub Actions에서 Run workflow를 다시 실행하세요."
                ) from exc

            delay = _DART_RETRY_DELAYS[attempt - 1]
            print(
                "Open DART 연결이 지연되어 "
                f"{delay}초 후 다시 시도합니다 "
                f"({attempt}/{_DART_CONNECTION_ATTEMPTS})...",
                file=sys.stderr,
            )
            time.sleep(delay)


def _looks_like_missing_xbrl(exc: Exception) -> bool:
    message = str(exc).lower()
    return (
        isinstance(exc, FileNotFoundError)
        or exc.__class__.__name__ == "NoDataReceived"
        or ("xbrl" in message and "not found" in message)
        or "no data received" in message
        or "조회된 데이타가 없습니다" in message
        or "조회된 데이터가 없습니다" in message
        or "파일이 존재하지 않습니다" in message
        or "target does not exist" in message
    )


def load_xbrl(receipt_number: str, download_dir: Path, dart_module: Any) -> Any:
    """Download and load one filing's XBRL through dart-fss."""

    try:
        xbrl_path = dart_module.api.finance.download_xbrl(
            path=str(download_dir),
            rcept_no=receipt_number,
        )
    except Exception as exc:
        if _looks_like_missing_xbrl(exc):
            raise MissingXbrlError(
                f"접수번호 {receipt_number} 공시에서 XBRL 파일을 찾을 수 없습니다."
            ) from exc
        raise FinancialExtractionError(f"XBRL 다운로드 중 오류가 발생했습니다: {exc}") from exc

    if not xbrl_path or not Path(xbrl_path).is_file():
        raise MissingXbrlError(
            f"접수번호 {receipt_number} 공시에서 XBRL 파일을 찾을 수 없습니다."
        )

    try:
        xbrl = dart_module.xbrl.get_xbrl_from_file(str(xbrl_path))
    except Exception as exc:
        if _looks_like_missing_xbrl(exc):
            raise MissingXbrlError(
                f"접수번호 {receipt_number} 공시의 XBRL 파일을 불러올 수 없습니다."
            ) from exc
        raise FinancialExtractionError(f"XBRL 로드 중 오류가 발생했습니다: {exc}") from exc

    is_empty = getattr(xbrl, "is_empty", False) if xbrl is not None else True
    if callable(is_empty):
        is_empty = is_empty()
    if xbrl is None or is_empty:
        raise MissingXbrlError(
            f"접수번호 {receipt_number} 공시에 사용 가능한 XBRL 데이터가 없습니다."
        )
    return xbrl


def _first_table(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
        return next((item for item in value if item is not None), None)
    return value


def _table_to_dataframe(table: Any, *, separate: bool) -> pd.DataFrame | None:
    if table is None:
        return None
    frame = table.to_DataFrame(
        lang="ko",
        label="Separate" if separate else "Consolidated",
        show_abstract=True,
        show_class=False,
        show_concept=False,
        separator=False,
    )
    if frame is None or frame.empty:
        return None
    simplified = simplify_statement_dataframe(frame)
    return simplified if not simplified.empty else None


def _extract_for_scope(
    xbrl: Any,
    *,
    separate: bool,
) -> tuple[dict[str, pd.DataFrame], list[str]]:
    statements: dict[str, pd.DataFrame] = {}
    missing: list[str] = []

    for statement_name, method_name in STATEMENT_SPECS:
        method = getattr(xbrl, method_name)
        table = _first_table(method(separate=separate))
        frame = _table_to_dataframe(table, separate=separate)
        if frame is None:
            missing.append(statement_name)
        else:
            statements[statement_name] = frame

    return statements, missing


def _scope_attempts(xbrl: Any, scope: str) -> list[tuple[str, bool]]:
    if scope == "consolidated":
        return [("consolidated", False)]
    if scope == "separate":
        return [("separate", True)]

    has_consolidated = bool(xbrl.exist_consolidated())
    if has_consolidated:
        return [("consolidated", False), ("separate", True)]
    return [("separate", True)]


def extract_statements(
    xbrl: Any,
    scope: str = "auto",
) -> tuple[dict[str, pd.DataFrame], str]:
    """Extract the required three statements and return their selected scope."""

    if scope not in {"auto", "consolidated", "separate"}:
        raise ValueError(f"지원하지 않는 재무제표 범위입니다: {scope}")

    failures: list[str] = []
    for scope_name, separate in _scope_attempts(xbrl, scope):
        statements, missing = _extract_for_scope(xbrl, separate=separate)
        if not missing:
            return statements, scope_name
        failures.append(f"{SCOPE_LABELS[scope_name]}: {', '.join(missing)}")

    details = "; ".join(failures)
    raise MissingStatementError(
        "필수 재무제표를 모두 찾을 수 없습니다. "
        f"누락된 재무제표 ({details}). "
        "재무상태표, 손익계산서(또는 포괄손익계산서), 현금흐름표가 모두 필요합니다."
    )


def _column_part_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = " ".join(str(value).split())
    if not text or text.lower().startswith("unnamed:"):
        return ""
    return text


def _flatten_column(column: Any) -> str:
    parts = column if isinstance(column, tuple) else (column,)
    cleaned: list[str] = []
    for part in parts:
        nested_parts = part if isinstance(part, tuple) else (part,)
        for nested_part in nested_parts:
            text = _column_part_text(nested_part)
            if text and (not cleaned or cleaned[-1] != text):
                cleaned.append(text)
    return " | ".join(cleaned) or "열"


def _deduplicate_headers(headers: Sequence[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique: list[str] = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        occurrence = counts[header]
        unique.append(header if occurrence == 1 else f"{header} ({occurrence})")
    return unique


def flatten_dataframe_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Copy a DataFrame and replace any multi-level columns with one header row."""

    flattened = frame.copy()
    headers = [_flatten_column(column) for column in flattened.columns.to_list()]
    flattened.columns = _deduplicate_headers(headers)
    return flattened


_DATE_PATTERN = re.compile(
    r"(?<!\d)((?:19|20)\d{2})[-./년]\s*(\d{1,2})[-./월]\s*(\d{1,2})(?:일)?(?!\d)"
)
_COMPACT_DATE_PATTERN = re.compile(r"(?<!\d)((?:19|20)\d{6})(?!\d)")
_ITEM_COLUMN_NAMES = {
    "labelko",
    "계정",
    "계정과목",
    "계정명",
    "과목",
    "항목",
}
_METADATA_COLUMN_NAMES = {
    "concept",
    "conceptid",
    "labelen",
    "class",
    "주석",
    "note",
    "notes",
}


def _column_parts(column: Any) -> list[str]:
    values = column if isinstance(column, tuple) else (column,)
    parts: list[str] = []
    for value in values:
        nested_values = value if isinstance(value, tuple) else (value,)
        for nested_value in nested_values:
            text = _column_part_text(nested_value)
            if text and text not in parts:
                parts.append(text)
    return parts


def _normalized_header_part(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value).lower()


def _item_column_score(column: Any) -> int:
    normalized = [_normalized_header_part(part) for part in _column_parts(column)]
    if "labelko" in normalized:
        return 100
    if any(part in {"과목", "항목", "계정과목", "계정명"} for part in normalized):
        return 90
    if "계정" in normalized:
        return 80
    return 0


def _is_metadata_column(column: Any) -> bool:
    normalized = [_normalized_header_part(part) for part in _column_parts(column)]
    return any(
        part in _METADATA_COLUMN_NAMES
        or part.startswith("class")
        or part.startswith("conceptid")
        or part.startswith("주석")
        or part.startswith("note")
        for part in normalized
    )


def _as_excel_number(value: Any) -> int | float | Any:
    if value is None or isinstance(value, bool):
        return "" if value is None else value
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, numbers.Number):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    text = " ".join(str(value).replace("\u00a0", " ").split())
    if not text or text in {"-", "–", "—"}:
        return ""

    negative = text.startswith("(") and text.endswith(")")
    numeric_text = text[1:-1] if negative else text
    numeric_text = numeric_text.replace(",", "")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", numeric_text):
        number: int | float
        number = float(numeric_text) if "." in numeric_text else int(numeric_text)
        return -number if negative else number
    return text


def _series_has_numeric_value(series: pd.Series) -> bool:
    return any(
        isinstance(converted := _as_excel_number(value), numbers.Number)
        and not isinstance(converted, bool)
        for value in series
    )


def _period_header(column: Any) -> str:
    parts = _column_parts(column)
    joined = " ".join(parts)
    dates = [
        f"{year}-{int(month):02d}-{int(day):02d}"
        for year, month, day in _DATE_PATTERN.findall(joined)
    ]
    dates.extend(
        f"{date[:4]}-{date[4:6]}-{date[6:]}"
        for date in _COMPACT_DATE_PATTERN.findall(joined)
    )
    dates = list(dict.fromkeys(dates))
    if dates:
        return dates[0] if len(dates) == 1 else f"{dates[0]} ~ {dates[-1]}"

    preferred_markers = ("당기", "전기", "분기", "반기", "누적", "기말", "연도")
    for part in reversed(parts):
        compact = _normalized_header_part(part)
        if any(marker in compact for marker in preferred_markers):
            return part

    for part in reversed(parts):
        compact = _normalized_header_part(part)
        if (
            compact not in _METADATA_COLUMN_NAMES
            and compact not in _ITEM_COLUMN_NAMES
            and "unit:" not in part.lower()
            and "재무상태표" not in compact
            and "손익계산서" not in compact
            and "현금흐름표" not in compact
        ):
            return part
    return "금액"


def _clean_item_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = " ".join(str(value).replace("\u00a0", " ").split())
    return re.sub(r"\s*\[abstract\]\s*$", "", text, flags=re.IGNORECASE)


def _is_blank_cell(value: Any) -> bool:
    if value is None or (isinstance(value, str) and value == ""):
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _is_document_noise_item(value: Any) -> bool:
    text = _clean_item_value(value)
    compact = re.sub(r"\s+", "", text).casefold()
    if not compact:
        return False
    if compact in {
        "#name?",
        "재무상태표",
        "손익계산서",
        "포괄손익계산서",
        "현금흐름표",
    }:
        return True
    if any(
        marker in compact
        for marker in (
            "회사명:",
            "감사받지아니한재무제표",
            "별첨재무제표",
            "별첨주석",
            "첨부된주석",
        )
    ):
        return True
    if re.match(r"^제(?:\d+)?기", compact) and re.search(
        r"(?:19|20)\d{2}년",
        compact,
    ):
        return True
    if re.match(r"^제\((?:당|전)\)기$", compact):
        return True
    if re.match(r"^\(?단위:", compact):
        return True
    if "과목" in compact and any(marker in compact for marker in ("당", "전")):
        return True
    return False


def _merge_orphan_amount_rows(frame: pd.DataFrame) -> pd.DataFrame:
    """Attach an amount-only continuation row to a safe adjacent item row."""

    if frame.empty or len(frame) < 2:
        return frame

    merged = frame.copy()
    drop_indices: list[int] = []
    for row_index in range(len(merged)):
        if _clean_item_value(merged.iat[row_index, 0]):
            continue
        orphan_values = list(merged.iloc[row_index, 1:])
        if not any(not _is_blank_cell(value) for value in orphan_values):
            drop_indices.append(row_index)
            continue

        candidate_indices = [
            index
            for index in (row_index + 1, row_index - 1)
            if 0 <= index < len(merged)
            and _clean_item_value(merged.iat[index, 0])
        ]
        compatible = [
            index
            for index in candidate_indices
            if all(
                _is_blank_cell(orphan)
                or _is_blank_cell(merged.iat[index, column_index])
                for column_index, orphan in enumerate(orphan_values, start=1)
            )
        ]
        if not compatible:
            continue

        target_index = compatible[0]
        for column_index, orphan in enumerate(orphan_values, start=1):
            if (
                not _is_blank_cell(orphan)
                and _is_blank_cell(merged.iat[target_index, column_index])
            ):
                merged.iat[target_index, column_index] = orphan
        drop_indices.append(row_index)

    if drop_indices:
        merged = merged.drop(index=drop_indices)
    return merged.reset_index(drop=True)


def simplify_statement_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    """Keep one Korean item column and only period columns containing amounts."""

    if frame is None or frame.empty:
        return pd.DataFrame()

    columns = frame.columns.to_list()
    item_index = max(
        range(len(columns)),
        key=lambda index: _item_column_score(columns[index]),
        default=0,
    )
    if _item_column_score(columns[item_index]) == 0:
        item_index = 0

    amount_indices = [
        index
        for index, column in enumerate(columns)
        if index != item_index
        and not _is_metadata_column(column)
        and _series_has_numeric_value(frame.iloc[:, index])
    ]
    if not amount_indices:
        return pd.DataFrame()

    output_columns = ["항목"] + [
        _period_header(columns[index]) for index in amount_indices
    ]
    output_columns = _deduplicate_headers(output_columns)

    series = [frame.iloc[:, item_index].map(_clean_item_value)]
    series.extend(
        frame.iloc[:, index].map(_as_excel_number) for index in amount_indices
    )
    simplified = pd.concat(series, axis=1)
    simplified.columns = output_columns

    populated_rows = simplified.apply(
        lambda row: any(value not in {"", None} for value in row),
        axis=1,
    )
    simplified = simplified.loc[populated_rows].reset_index(drop=True)
    simplified = simplified.loc[
        ~simplified.iloc[:, 0].map(_is_document_noise_item)
    ].reset_index(drop=True)
    return _merge_orphan_amount_rows(simplified)


_LEADING_ITEM_NUMBER_PATTERN = re.compile(
    r"^\s*(?:(?:[IVXLCDM]+|[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫ]+|\d+|[가-힣])"
    r"\s*[\.\)\-:]\s*|\(\s*(?:\d+|[가-힣])\s*\)\s*)+",
    flags=re.IGNORECASE,
)
_TRAILING_NOTE_REFERENCE_PATTERN = re.compile(
    r"\s*[\(\[]\s*(?:주석|notes?)\s*[\d,\-~·ㆍ\s]+\s*[\)\]]\s*$",
    flags=re.IGNORECASE,
)


def parse_disclosure_inputs(value: str) -> list[str]:
    """Split comma/newline-separated filings and reject duplicate receipts."""

    disclosures = [part.strip() for part in re.split(r"[,\r\n]+", value) if part.strip()]
    if not disclosures:
        raise FinancialExtractionError(
            "DART 공시 URL 또는 14자리 접수번호를 하나 이상 입력하세요."
        )

    seen_receipts: set[str] = set()
    for disclosure in disclosures:
        receipt_number = parse_disclosure_reference(disclosure).receipt_number
        if receipt_number in seen_receipts:
            raise FinancialExtractionError(
                f"같은 접수번호가 두 번 입력되었습니다: {receipt_number}"
            )
        seen_receipts.add(receipt_number)
    return disclosures


def _normalized_item_key(value: Any) -> str:
    label = unicodedata.normalize("NFKC", _clean_item_value(value))
    previous = None
    while label != previous:
        previous = label
        label = _LEADING_ITEM_NUMBER_PATTERN.sub("", label)
    label = _TRAILING_NOTE_REFERENCE_PATTERN.sub("", label)
    return re.sub(r"\s+", "", label).casefold()


def _period_dates(header: str) -> list[str]:
    dates = [
        f"{year}-{int(month):02d}-{int(day):02d}"
        for year, month, day in _DATE_PATTERN.findall(header)
    ]
    dates.extend(
        f"{date[:4]}-{date[4:6]}-{date[6:]}"
        for date in _COMPACT_DATE_PATTERN.findall(header)
    )
    return list(dict.fromkeys(dates))


def _qualified_period_header(header: str, metadata: FilingMetadata) -> str:
    if _period_dates(header):
        return header
    return f"{metadata.year_month} | {header}"


def _period_sort_key(header: str) -> tuple[str, str]:
    dates = _period_dates(header)
    return (max(dates) if dates else "", header)


def _described_period_headers(frame: pd.DataFrame) -> list[str]:
    if frame is None or frame.empty:
        return []

    columns = frame.columns.to_list()
    item_index = max(
        range(len(columns)),
        key=lambda index: _item_column_score(columns[index]),
        default=0,
    )
    if _item_column_score(columns[item_index]) == 0:
        item_index = 0

    periods: list[str] = []
    for value in frame.iloc[:, item_index]:
        text = _clean_item_value(value)
        compact = re.sub(r"\s+", "", text)
        dates = _period_dates(text)
        if (
            dates
            and re.match(r"^제(?:\d+)?기", compact)
            and any(marker in compact for marker in ("현재", "부터", "까지"))
        ):
            period = dates[0] if len(dates) == 1 else f"{dates[0]} ~ {dates[-1]}"
            if period not in periods:
                periods.append(period)

    return sorted(periods, key=_period_sort_key, reverse=True)


def _resolved_period_headers(
    raw_frame: pd.DataFrame,
    simplified: pd.DataFrame,
    metadata: FilingMetadata,
) -> list[str]:
    described_periods = _described_period_headers(raw_frame)
    resolved: list[str] = []
    for column in simplified.columns[1:]:
        header = str(column)
        if _period_dates(header):
            resolved.append(header)
            continue

        compact = _normalized_header_part(header)
        if "당" in compact and described_periods:
            resolved.append(described_periods[0])
        elif "전" in compact and len(described_periods) >= 2:
            resolved.append(described_periods[1])
        else:
            resolved.append(_qualified_period_header(header, metadata))
    return _deduplicate_headers(resolved)


def _period_duration_days(header: str) -> int:
    dates = _period_dates(header)
    if len(dates) < 2:
        return 0
    return (date.fromisoformat(dates[-1]) - date.fromisoformat(dates[0])).days


def _primary_period_index(
    headers: Sequence[str],
    metadata: FilingMetadata,
) -> int:
    dated = [
        (index, _period_dates(header))
        for index, header in enumerate(headers)
        if _period_dates(header)
    ]
    if dated:
        latest_end = max(max(dates) for _, dates in dated)
        candidates = [
            index for index, dates in dated if max(dates) == latest_end
        ]
        if len(candidates) == 1:
            return candidates[0]
        if metadata.report_type == "분기보고서":
            return min(candidates, key=lambda index: _period_duration_days(headers[index]))
        return max(candidates, key=lambda index: _period_duration_days(headers[index]))

    current = [
        index
        for index, header in enumerate(headers)
        if "당" in _normalized_header_part(header)
        and "전" not in _normalized_header_part(header)
    ]
    return current[0] if current else 0


def _display_period_label(header: str, metadata: FilingMetadata) -> str:
    dates = _period_dates(header)
    if dates:
        end_date = date.fromisoformat(max(dates))
        year = end_date.year
        month = end_date.month
    else:
        filing_year, filing_month = (
            int(part) for part in metadata.year_month.split(".", maxsplit=1)
        )
        year = (
            filing_year - 1
            if metadata.report_type in {"사업보고서", "감사보고서"}
            else filing_year
        )
        month = {
            "분기보고서": 3 if filing_month <= 6 else 9,
            "반기보고서": 6,
        }.get(metadata.report_type, 12)

    if metadata.report_type in {"사업보고서", "감사보고서"}:
        return f"{year}년"
    if metadata.report_type == "반기보고서":
        return f"{year}년 반기"
    quarter = max(1, min(4, (month - 1) // 3 + 1))
    return f"{year}년 {quarter}분기"


def _has_primary_amount_or_core_section(row: pd.Series) -> bool:
    label = re.sub(r"\s+", "", _clean_item_value(row.iloc[0]))
    if not label:
        return False
    if any(not _is_blank_cell(value) for value in row.iloc[1:]):
        return True
    return label in {"자산", "부채", "자본"}


def merge_statement_frames(
    filings: Sequence[tuple[FilingMetadata, pd.DataFrame]],
    *,
    primary_period_only: bool = False,
) -> pd.DataFrame:
    """Merge one statement across filings without fuzzy-matching changed items."""

    rows: list[dict[str, Any]] = []
    period_headers: set[str] = set()

    for metadata, raw_frame in sorted(
        filings,
        key=lambda item: item[0].year_month,
    ):
        frame = simplify_statement_dataframe(raw_frame)
        if frame.empty:
            raise MissingStatementError(
                "연도별 병합 중 항목과 기간별 금액으로 정리할 수 없는 "
                "재무제표가 발견되었습니다."
            )

        resolved_headers = _resolved_period_headers(raw_frame, frame, metadata)
        frame.columns = ["항목", *resolved_headers]
        if primary_period_only:
            primary_index = _primary_period_index(resolved_headers, metadata)
            display_header = _display_period_label(
                resolved_headers[primary_index],
                metadata,
            )
            frame = frame.iloc[:, [0, primary_index + 1]].copy()
            frame.columns = ["항목", display_header]
            frame = frame.loc[
                frame.apply(_has_primary_amount_or_core_section, axis=1)
            ].reset_index(drop=True)

        current_labels = [_clean_item_value(value) for value in frame.iloc[:, 0]]
        current_keys = [_normalized_item_key(value) for value in current_labels]
        master_keys = [str(row["key"]) for row in rows]
        matcher = SequenceMatcher(
            None,
            master_keys,
            current_keys,
            autojunk=False,
        )

        aligned_rows: list[dict[str, Any]] = []
        current_mapping: dict[int, dict[str, Any]] = {}
        for tag, old_start, old_end, new_start, new_end in matcher.get_opcodes():
            if tag == "equal":
                for offset, current_index in enumerate(range(new_start, new_end)):
                    row = rows[old_start + offset]
                    row["label"] = current_labels[current_index]
                    aligned_rows.append(row)
                    current_mapping[current_index] = row
                continue

            if tag in {"replace", "insert"}:
                for current_index in range(new_start, new_end):
                    row = {
                        "key": current_keys[current_index],
                        "label": current_labels[current_index],
                        "values": {},
                    }
                    aligned_rows.append(row)
                    current_mapping[current_index] = row

            if tag in {"replace", "delete"}:
                aligned_rows.extend(rows[old_start:old_end])

        rows = aligned_rows
        current_periods = [str(column) for column in frame.columns[1:]]

        for period in current_periods:
            if period in period_headers:
                for row in rows:
                    row["values"].pop(period, None)
            period_headers.add(period)

        for current_index, row in current_mapping.items():
            for column_index, period in enumerate(current_periods, start=1):
                row["values"][period] = _as_excel_number(
                    frame.iloc[current_index, column_index]
                )

    ordered_periods = sorted(period_headers, key=_period_sort_key, reverse=True)
    output_rows = [
        [row["label"]] + [row["values"].get(period, "") for period in ordered_periods]
        for row in rows
    ]
    return pd.DataFrame(output_rows, columns=["항목", *ordered_periods])


def merge_filing_statements(
    filings: Sequence[FilingExtraction],
) -> dict[str, pd.DataFrame]:
    """Merge exactly the required three statements across filings."""

    return {
        statement_name: merge_statement_frames(
            [
                (filing.metadata, filing.statements[statement_name])
                for filing in filings
            ],
            primary_period_only=True,
        )
        for statement_name in STATEMENT_NAMES
    }


def _normalized_company_name(value: str) -> str:
    name = unicodedata.normalize("NFKC", value)
    name = re.sub(r"(?:주식회사|\(주\)|㈜)", "", name)
    return re.sub(r"\s+", "", name).casefold()


def _validate_merge_compatibility(
    filings: Sequence[FilingExtraction],
) -> None:
    companies = {
        _normalized_company_name(filing.metadata.company_name) for filing in filings
    }
    if len(companies) != 1:
        names = ", ".join(dict.fromkeys(f.metadata.company_name for f in filings))
        raise FinancialExtractionError(
            f"연도별 병합에는 같은 회사의 공시만 입력할 수 있습니다: {names}"
        )

    report_types = {filing.metadata.report_type for filing in filings}
    if len(report_types) != 1:
        names = ", ".join(sorted(report_types))
        raise FinancialExtractionError(
            f"연도별 병합에는 같은 보고서 종류만 입력할 수 있습니다: {names}"
        )

    scopes = {filing.selected_scope for filing in filings}
    if len(scopes) != 1:
        names = ", ".join(SCOPE_LABELS[scope] for scope in sorted(scopes))
        raise FinancialExtractionError(
            "연결과 별도 재무제표를 한 파일에 섞을 수 없습니다. "
            f"각 공시에서 선택된 범위: {names}"
        )


def _merged_year_range(statements: Mapping[str, pd.DataFrame]) -> str:
    years: set[str] = set()
    for frame in statements.values():
        for column in frame.columns[1:]:
            dates = _period_dates(str(column))
            years.update(date[:4] for date in dates)
            years.update(re.findall(r"(?:19|20)\d{2}", str(column)))
    if not years:
        raise FinancialExtractionError(
            "병합된 재무제표의 기간에서 연도를 확인할 수 없습니다."
        )
    ordered = sorted(years)
    return ordered[0] if len(ordered) == 1 else f"{ordered[0]}-{ordered[-1]}"


def _display_width(value: Any) -> int:
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
        for character in text
    )


def format_worksheet(worksheet: Any) -> None:
    """Apply the small, deterministic formatting set required for outputs."""

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(min_row=2):
        has_number = False
        for cell in row:
            value = cell.value
            if isinstance(value, numbers.Number) and not isinstance(value, bool):
                if isinstance(value, float) and not math.isfinite(value):
                    continue
                cell.number_format = "#,##0;[Red](#,##0);-"
                cell.alignment = Alignment(horizontal="right")
                has_number = True
            elif cell.column == 1:
                cell.alignment = Alignment(horizontal="left")
        if not has_number and row[0].value:
            row[0].font = Font(bold=True)

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        content_width = max((_display_width(cell.value) for cell in column_cells), default=0)
        minimum_width = 24 if column_index == 1 else 16
        maximum_width = 45 if column_index == 1 else 24
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(content_width + 2, minimum_width),
            maximum_width,
        )


def _write_workbook(path: Path, sheets: Mapping[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            format_worksheet(writer.book[sheet_name])


def export_statements(
    statements: Mapping[str, pd.DataFrame],
    *,
    metadata: FilingMetadata,
    output_mode: str,
    output_dir: Path,
) -> list[Path]:
    """Write one three-sheet workbook or three one-sheet workbooks."""

    missing = [name for name in STATEMENT_NAMES if name not in statements]
    extras = [name for name in statements if name not in STATEMENT_NAMES]
    if missing or extras:
        raise ValueError(
            "내보낼 재무제표는 정확히 재무상태표, 손익계산서, 현금흐름표여야 합니다."
        )

    ordered_statements = {
        name: simplify_statement_dataframe(statements[name])
        for name in STATEMENT_NAMES
    }
    empty = [name for name, frame in ordered_statements.items() if frame.empty]
    if empty:
        raise ValueError(
            "항목과 기간별 금액으로 정리할 수 없는 재무제표가 있습니다: "
            + ", ".join(empty)
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    filename_stem = metadata.filename_stem

    if output_mode == "single":
        workbook_path = output_dir / f"{filename_stem}.xlsx"
        _write_workbook(workbook_path, ordered_statements)
        return [workbook_path]

    if output_mode == "separate":
        paths: list[Path] = []
        for statement_name, frame in ordered_statements.items():
            workbook_path = output_dir / f"{filename_stem}_{statement_name}.xlsx"
            _write_workbook(workbook_path, {statement_name: frame})
            paths.append(workbook_path)
        return paths

    raise ValueError(f"지원하지 않는 출력 방식입니다: {output_mode}")


def _extract_filing(
    disclosure: str,
    *,
    scope: str,
    dart: Any,
) -> FilingExtraction:
    """Extract one filing without configuring credentials or exporting files."""

    reference = parse_disclosure_reference(disclosure)
    receipt_number = reference.receipt_number
    metadata = resolve_filing_metadata(receipt_number, dart)

    with tempfile.TemporaryDirectory(prefix="dart-financials-") as temp_dir:
        temp_path = Path(temp_dir)
        try:
            xbrl = load_xbrl(receipt_number, temp_path, dart)
        except MissingXbrlError as xbrl_error:
            try:
                document_number = reference.document_number or resolve_document_number(
                    receipt_number,
                    dart,
                )
                pdf_path = download_filing_pdf(
                    receipt_number,
                    document_number,
                    temp_path,
                    dart,
                )
                statements, selected_scope = extract_pdf_statements(pdf_path, scope)
            except PdfExtractionError as pdf_error:
                raise FinancialExtractionError(
                    f"{xbrl_error} PDF 대체 추출도 실패했습니다: {pdf_error}"
                ) from pdf_error
            source = "PDF"
        else:
            statements, selected_scope = extract_statements(xbrl, scope)
            source = "XBRL"

    return FilingExtraction(
        metadata=metadata,
        statements=statements,
        selected_scope=selected_scope,
        receipt_number=receipt_number,
        source=source,
    )


def _configured_dart(
    *,
    environ: Mapping[str, str] | None,
    dart_module: Any | None,
) -> Any:
    api_key = require_api_key(environ)
    dart = _import_dart_fss() if dart_module is None else dart_module
    _set_dart_api_key_with_retry(dart, api_key)
    return dart


def run(
    disclosure: str,
    *,
    scope: str = "auto",
    output_mode: str = "single",
    output_dir: Path = Path("output"),
    environ: Mapping[str, str] | None = None,
    dart_module: Any | None = None,
) -> tuple[list[Path], str, str, str]:
    """Run one-filing extraction and keep the original public return shape."""

    dart = _configured_dart(environ=environ, dart_module=dart_module)
    filing = _extract_filing(disclosure, scope=scope, dart=dart)
    paths = export_statements(
        filing.statements,
        metadata=filing.metadata,
        output_mode=output_mode,
        output_dir=output_dir,
    )

    return (
        paths,
        filing.selected_scope,
        filing.receipt_number,
        filing.source,
    )


def run_many(
    disclosure_input: str,
    *,
    scope: str = "auto",
    output_mode: str = "single",
    output_dir: Path = Path("output"),
    environ: Mapping[str, str] | None = None,
    dart_module: Any | None = None,
) -> tuple[list[Path], list[FilingExtraction]]:
    """Extract one or more filings, merging multiple years into one output."""

    disclosures = parse_disclosure_inputs(disclosure_input)
    dart = _configured_dart(environ=environ, dart_module=dart_module)
    filings = [
        _extract_filing(disclosure, scope=scope, dart=dart)
        for disclosure in disclosures
    ]

    if len(filings) == 1:
        statements = filings[0].statements
        output_metadata = filings[0].metadata
    else:
        _validate_merge_compatibility(filings)
        statements = merge_filing_statements(filings)
        first_metadata = filings[0].metadata
        output_metadata = FilingMetadata(
            company_name=first_metadata.company_name,
            report_type=first_metadata.report_type,
            year_month=_merged_year_range(statements),
        )

    paths = export_statements(
        statements,
        metadata=output_metadata,
        output_mode=output_mode,
        output_dir=output_dir,
    )
    return paths, filings


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "DART 공시 XBRL에서 재무제표 3종을 추출하고, "
            "XBRL이 없으면 공시 PDF를 사용합니다."
        )
    )
    parser.add_argument(
        "disclosure",
        help=(
            "DART 공시 URL 또는 14자리 접수번호. 여러 연도는 쉼표 또는 줄바꿈으로 "
            "구분합니다."
        ),
    )
    parser.add_argument(
        "--scope",
        choices=("auto", "consolidated", "separate"),
        default="auto",
        help="auto는 연결 3종을 우선하고 불가능하면 별도 3종으로 재시도합니다.",
    )
    parser.add_argument(
        "--output-mode",
        choices=("single", "separate"),
        default="single",
        help="single은 3개 시트 통합 파일, separate는 개별 파일 3개입니다.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output"),
        help="Excel 파일을 저장할 폴더(기본값: output)",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_argument_parser().parse_args(argv)
    try:
        paths, filings = run_many(
            args.disclosure,
            scope=args.scope,
            output_mode=args.output_mode,
            output_dir=args.output_dir,
        )
    except (
        InvalidDisclosureReference,
        FinancialExtractionError,
        ValueError,
    ) as exc:
        print(f"오류: {exc}", file=sys.stderr)
        return 1

    for filing in filings:
        print(
            f"접수번호: {filing.receipt_number} | "
            f"추출 원본: {filing.source} | "
            f"재무제표 범위: {SCOPE_LABELS[filing.selected_scope]}"
        )
    for path in paths:
        print(f"생성 완료: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
