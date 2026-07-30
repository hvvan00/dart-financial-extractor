import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from extract_financials import (
    FilingExtraction,
    FilingMetadata,
    FinancialExtractionError,
    MissingApiKeyError,
    MissingStatementError,
    MissingXbrlError,
    STATEMENT_NAMES,
    _set_dart_api_key_with_retry,
    export_statements,
    extract_statements,
    flatten_dataframe_columns,
    load_xbrl,
    merge_statement_frames,
    parse_disclosure_inputs,
    parse_filing_metadata_html,
    require_api_key,
    resolve_filing_metadata,
    run,
    run_many,
    simplify_statement_dataframe,
)


class FakeTable:
    def __init__(self, value, separate):
        self.value = value
        self.separate = separate

    def to_DataFrame(self, **kwargs):
        expected_options = {
            "lang": "ko",
            "label": "Separate" if self.separate else "Consolidated",
            "show_abstract": True,
            "show_class": False,
            "show_concept": False,
            "separator": False,
        }
        if kwargs != expected_options:
            raise AssertionError(f"unexpected XBRL table options: {kwargs}")
        if kwargs.get("separator") is not False:
            raise AssertionError("numeric extraction must disable string separators")
        definition = "연결 재무제표 (Unit: KRW)"
        return pd.DataFrame(
            [[self.value, 1_234_567.0]],
            columns=pd.MultiIndex.from_tuples(
                [
                    (definition, "label_ko"),
                    (definition, "[2025-12-31]연결재무제표"),
                ]
            ),
        )


class FakeXbrl:
    def __init__(self, consolidated=True, missing_consolidated=()):
        self.consolidated = consolidated
        self.missing_consolidated = set(missing_consolidated)
        self.calls = []

    def exist_consolidated(self):
        return self.consolidated

    def _result(self, name, separate):
        self.calls.append((name, separate))
        if not separate and name in self.missing_consolidated:
            return None
        return [FakeTable(f"{name}-{'별도' if separate else '연결'}", separate)]

    def get_financial_statement(self, separate=False):
        return self._result("재무상태표", separate)

    def get_income_statement(self, separate=False):
        return self._result("손익계산서", separate)

    def get_cash_flows(self, separate=False):
        return self._result("현금흐름표", separate)


class ApiKeyTests(unittest.TestCase):
    def test_missing_api_key_has_clear_error(self):
        with self.assertRaisesRegex(MissingApiKeyError, "DART_API_KEY"):
            require_api_key({})

    def test_blank_api_key_has_clear_error(self):
        with self.assertRaisesRegex(MissingApiKeyError, "DART_API_KEY"):
            require_api_key({"DART_API_KEY": "  "})

    def test_reads_api_key_from_named_environment_variable(self):
        self.assertEqual(
            require_api_key({"DART_API_KEY": " secret-value "}),
            "secret-value",
        )

    def test_transient_timeout_is_retried_until_authentication_succeeds(self):
        outcomes = iter(
            [
                TimeoutError("connection timed out"),
                TimeoutError("connection timed out"),
                None,
            ]
        )
        calls = []

        def set_api_key(**kwargs):
            calls.append(kwargs)
            outcome = next(outcomes)
            if outcome is not None:
                raise outcome

        dart = SimpleNamespace(set_api_key=set_api_key)
        with patch("extract_financials.time.sleep") as sleep:
            _set_dart_api_key_with_retry(dart, "secret-value")

        self.assertEqual(len(calls), 3)
        self.assertEqual([call.args[0] for call in sleep.call_args_list], [5, 15])

    def test_repeated_timeout_has_clear_rerun_error_without_exposing_key(self):
        calls = []

        def set_api_key(**kwargs):
            calls.append(kwargs)
            raise TimeoutError("connection timed out")

        dart = SimpleNamespace(set_api_key=set_api_key)
        with (
            patch("extract_financials.time.sleep"),
            self.assertRaises(FinancialExtractionError) as raised,
        ):
            _set_dart_api_key_with_retry(dart, "do-not-print-this-key")

        self.assertEqual(len(calls), 3)
        self.assertIn("3번", str(raised.exception))
        self.assertIn("다시 실행", str(raised.exception))
        self.assertNotIn("do-not-print-this-key", str(raised.exception))

    def test_non_network_authentication_error_is_not_retried(self):
        calls = []

        def set_api_key(**kwargs):
            calls.append(kwargs)
            raise RuntimeError("invalid API key")

        dart = SimpleNamespace(set_api_key=set_api_key)
        with (
            patch("extract_financials.time.sleep") as sleep,
            self.assertRaisesRegex(FinancialExtractionError, "DART_API_KEY"),
        ):
            _set_dart_api_key_with_retry(dart, "invalid-key")

        self.assertEqual(len(calls), 1)
        sleep.assert_not_called()


class FilingMetadataTests(unittest.TestCase):
    def test_parses_dart_slash_title(self):
        metadata = parse_filing_metadata_html(
            "<html><title>펀진/사업보고서/2026.03.31</title></html>"
        )

        self.assertEqual(metadata.company_name, "펀진")
        self.assertEqual(metadata.report_type, "사업보고서")
        self.assertEqual(metadata.year_month, "2026.03")
        self.assertEqual(metadata.filename_stem, "펀진_사업보고서_2026.03")

    def test_parses_bracket_pdf_style_title(self):
        metadata = parse_filing_metadata_html(
            "<title>[테스트 회사] 반기보고서(2025.08.14)</title>"
        )

        self.assertEqual(
            metadata.filename_stem,
            "테스트 회사_반기보고서_2025.08",
        )

    def test_parses_audit_report_title_with_year_and_month_only(self):
        metadata = parse_filing_metadata_html(
            "<title>[넥슨] 감사보고서(2025.12)</title>"
        )

        self.assertEqual(metadata.company_name, "넥슨")
        self.assertEqual(metadata.report_type, "감사보고서")
        self.assertEqual(metadata.year_month, "2025.12")
        self.assertEqual(metadata.filename_stem, "넥슨_감사보고서_2025.12")

    def test_invalid_filename_characters_are_replaced(self):
        metadata = FilingMetadata(
            company_name="테스트:회사",
            report_type="분기보고서",
            year_month="2025.11",
        )

        self.assertEqual(
            metadata.filename_stem,
            "테스트_회사_분기보고서_2025.11",
        )

    def test_missing_filename_metadata_has_clear_error(self):
        with self.assertRaisesRegex(
            FinancialExtractionError,
            "회사명, 보고서 종류, 연월",
        ):
            parse_filing_metadata_html("<title>DART 전자공시</title>")

    def test_resolves_metadata_from_disclosure_page(self):
        response = SimpleNamespace(
            text="<title>펀진/사업보고서/2026.03.31</title>",
            raise_for_status=lambda: None,
        )
        calls = []
        dart = SimpleNamespace(
            utils=SimpleNamespace(
                request=SimpleNamespace(
                    get=lambda **kwargs: calls.append(kwargs) or response
                )
            )
        )

        metadata = resolve_filing_metadata("20260331004320", dart)

        self.assertEqual(metadata.filename_stem, "펀진_사업보고서_2026.03")
        self.assertEqual(calls[0]["payload"], {"rcpNo": "20260331004320"})


class MultipleDisclosureInputTests(unittest.TestCase):
    def test_accepts_comma_and_newline_separated_disclosures(self):
        disclosures = parse_disclosure_inputs(
            "20240319000709,\n"
            "https://dart.fss.or.kr/dsaf001/main.do?rcpNo=20250319000710"
        )

        self.assertEqual(
            disclosures,
            [
                "20240319000709",
                "https://dart.fss.or.kr/dsaf001/main.do?rcpNo=20250319000710",
            ],
        )

    def test_duplicate_receipt_numbers_have_clear_error(self):
        with self.assertRaisesRegex(FinancialExtractionError, "두 번"):
            parse_disclosure_inputs(
                "20240319000709,\n"
                "https://dart.fss.or.kr/dsaf001/main.do?rcpNo=20240319000709"
            )


class XbrlLoadingTests(unittest.TestCase):
    def test_missing_xbrl_has_clear_error(self):
        finance = SimpleNamespace(
            download_xbrl=lambda **kwargs: (_ for _ in ()).throw(
                FileNotFoundError("XBRL File Not Found")
            )
        )
        dart = SimpleNamespace(api=SimpleNamespace(finance=finance))

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(MissingXbrlError, "XBRL"):
                load_xbrl("20240319000709", Path(temp_dir), dart)

    def test_open_dart_no_data_message_is_treated_as_missing_xbrl(self):
        finance = SimpleNamespace(
            download_xbrl=lambda **kwargs: (_ for _ in ()).throw(
                RuntimeError("013 조회된 데이터가 없습니다.")
            )
        )
        dart = SimpleNamespace(api=SimpleNamespace(finance=finance))

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(MissingXbrlError, "XBRL"):
                load_xbrl("20240319000709", Path(temp_dir), dart)

    def test_empty_xbrl_has_clear_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            xbrl_file = Path(temp_dir) / "filing.xbrl"
            xbrl_file.touch()
            finance = SimpleNamespace(download_xbrl=lambda **kwargs: str(xbrl_file))
            dart = SimpleNamespace(
                api=SimpleNamespace(finance=finance),
                xbrl=SimpleNamespace(
                    get_xbrl_from_file=lambda path: SimpleNamespace(is_empty=True)
                ),
            )

            with self.assertRaisesRegex(MissingXbrlError, "XBRL"):
                load_xbrl("20240319000709", Path(temp_dir), dart)


class StatementExtractionTests(unittest.TestCase):
    def test_auto_prefers_complete_consolidated_statements(self):
        xbrl = FakeXbrl(consolidated=True)

        statements, selected_scope = extract_statements(xbrl, "auto")

        self.assertEqual(selected_scope, "consolidated")
        self.assertEqual(tuple(statements), STATEMENT_NAMES)
        self.assertTrue(all(separate is False for _, separate in xbrl.calls))
        self.assertEqual(
            list(statements["재무상태표"].columns),
            ["항목", "2025-12-31"],
        )

    def test_auto_falls_back_to_complete_separate_statements(self):
        xbrl = FakeXbrl(
            consolidated=True,
            missing_consolidated={"현금흐름표"},
        )

        statements, selected_scope = extract_statements(xbrl, "auto")

        self.assertEqual(selected_scope, "separate")
        self.assertEqual(tuple(statements), STATEMENT_NAMES)
        self.assertIn(("현금흐름표", False), xbrl.calls)
        self.assertIn(("재무상태표", True), xbrl.calls)

    def test_auto_uses_separate_when_consolidated_does_not_exist(self):
        xbrl = FakeXbrl(consolidated=False)

        _, selected_scope = extract_statements(xbrl, "auto")

        self.assertEqual(selected_scope, "separate")
        self.assertTrue(all(separate is True for _, separate in xbrl.calls))

    def test_missing_statement_has_clear_error(self):
        xbrl = FakeXbrl(
            consolidated=True,
            missing_consolidated={"현금흐름표"},
        )
        xbrl.get_cash_flows = lambda separate=False: None

        with self.assertRaisesRegex(
            MissingStatementError,
            "현금흐름표",
        ):
            extract_statements(xbrl, "auto")


class ExcelOutputTests(unittest.TestCase):
    def setUp(self):
        definition = "연결 재무상태표 (Unit: KRW)"
        columns = pd.MultiIndex.from_tuples(
            [
                (definition, "concept_id"),
                (definition, "label_ko"),
                (definition, "label_en"),
                (definition, "class0"),
                (definition, "주석"),
                (definition, "[2025-12-31]연결재무제표"),
                (definition, "[2024-12-31]연결재무제표"),
            ]
        )
        self.statements = {
            name: pd.DataFrame(
                [
                    [
                        "dart_CashAndCashEquivalents",
                        "현금및현금성자산",
                        "Cash and cash equivalents",
                        "유동자산",
                        5,
                        1_234_567.0,
                        900_000.0,
                    ]
                ],
                columns=columns,
            )
            for name in STATEMENT_NAMES
        }

    def test_flattens_multi_level_headers_and_deduplicates(self):
        frame = pd.DataFrame(
            [[1, 2]],
            columns=pd.MultiIndex.from_tuples([("금액", "당기"), ("금액", "당기")]),
        )

        flattened = flatten_dataframe_columns(frame)

        self.assertEqual(
            flattened.columns.to_list(),
            ["금액 | 당기", "금액 | 당기 (2)"],
        )

    def test_simplifies_to_korean_item_and_period_amounts_only(self):
        definition = "연결 포괄손익계산서 (Unit: KRW)"
        frame = pd.DataFrame(
            [
                ["dart_AssetsAbstract", "자산 [abstract]", "Assets", "", None, None],
                [
                    "ifrs-full_CashAndCashEquivalents",
                    "현금및현금성자산",
                    "Cash and cash equivalents",
                    "5",
                    1_234_567.0,
                    900_000.0,
                ],
            ],
            columns=pd.MultiIndex.from_tuples(
                [
                    (definition, "concept_id"),
                    (definition, "label_ko"),
                    (definition, "label_en"),
                    (definition, "주석"),
                    (definition, "[2025-01-01,2025-12-31]연결재무제표"),
                    (definition, "[2024-01-01,2024-12-31]연결재무제표"),
                ]
            ),
        )

        simplified = simplify_statement_dataframe(frame)

        self.assertEqual(
            list(simplified.columns),
            [
                "항목",
                "2025-01-01 ~ 2025-12-31",
                "2024-01-01 ~ 2024-12-31",
            ],
        )
        self.assertEqual(simplified.iloc[0, 0], "자산")
        self.assertEqual(simplified.iloc[1, 0], "현금및현금성자산")
        self.assertEqual(simplified.iloc[1, 1], 1_234_567)
        self.assertNotIn("Cash and cash equivalents", simplified.to_string())

    def test_single_mode_has_exactly_three_formatted_sheets_and_numeric_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            paths = export_statements(
                self.statements,
                metadata=FilingMetadata("펀진", "사업보고서", "2026.03"),
                output_mode="single",
                output_dir=Path(temp_dir),
            )

            self.assertEqual(len(paths), 1)
            self.assertEqual(paths[0].name, "펀진_사업보고서_2026.03.xlsx")
            workbook = load_workbook(paths[0])
            self.assertEqual(workbook.sheetnames, list(STATEMENT_NAMES))
            for worksheet in workbook.worksheets:
                self.assertEqual(worksheet.freeze_panes, "A2")
                self.assertFalse(worksheet.sheet_view.showGridLines)
                self.assertTrue(worksheet["A1"].font.bold)
                self.assertEqual(
                    [cell.value for cell in worksheet[1]],
                    ["항목", "2025-12-31", "2024-12-31"],
                )
                self.assertEqual(worksheet["B2"].data_type, "n")
                self.assertEqual(worksheet["B2"].value, 1_234_567)
                self.assertIn(",", worksheet["B2"].number_format)
                self.assertNotIn("concept", " ".join(str(cell.value) for cell in worksheet[1]).lower())

    def test_formats_statement_hierarchy_with_category_bands_and_indentation(self):
        hierarchy_frame = pd.DataFrame(
            [
                ["자 산", "", ""],
                ["Ⅰ.유동자산", 100, 90],
                ["(1) 당좌자산", 100, 90],
                ["1. 현금및현금성자산", 110, 100],
                ["대손충당금", -10, -10],
                ["자 산 총 계", 100, 90],
            ],
            columns=["항목", "2025-12-31", "2024-12-31"],
        )
        statements = {name: hierarchy_frame.copy() for name in STATEMENT_NAMES}

        with tempfile.TemporaryDirectory() as temp_dir:
            path = export_statements(
                statements,
                metadata=FilingMetadata("테스트", "사업보고서", "2026.03"),
                output_mode="single",
                output_dir=Path(temp_dir),
            )[0]
            worksheet = load_workbook(path)["재무상태표"]

            self.assertEqual(worksheet["A1"].fill.fgColor.rgb[-6:], "006074")
            self.assertEqual(worksheet["A1"].font.color.rgb[-6:], "FFFFFF")
            self.assertEqual(worksheet["A2"].fill.fgColor.rgb[-6:], "006074")
            self.assertEqual(worksheet["A2"].font.color.rgb[-6:], "FFFFFF")
            self.assertEqual(worksheet["A3"].fill.fgColor.rgb[-6:], "008187")
            self.assertEqual(worksheet["A3"].font.color.rgb[-6:], "FFFFFF")
            self.assertNotIn("[Red]", worksheet["B3"].number_format)
            self.assertEqual(worksheet["A4"].fill.fgColor.rgb[-6:], "06A39F")
            self.assertEqual(worksheet["A4"].font.color.rgb[-6:], "FFFFFF")
            self.assertNotIn("[Red]", worksheet["B4"].number_format)
            self.assertEqual(worksheet["A5"].alignment.indent, 2)
            self.assertEqual(worksheet["A6"].alignment.indent, 3)
            self.assertTrue(worksheet["A7"].font.bold)
            self.assertEqual(worksheet["A7"].border.top.style, "medium")

    def test_separate_mode_creates_exactly_three_individual_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            paths = export_statements(
                self.statements,
                metadata=FilingMetadata("펀진", "사업보고서", "2026.03"),
                output_mode="separate",
                output_dir=Path(temp_dir),
            )

            self.assertEqual(len(paths), 3)
            self.assertEqual(len(list(Path(temp_dir).glob("*.xlsx"))), 3)
            for path, statement_name in zip(paths, STATEMENT_NAMES):
                self.assertEqual(
                    path.name,
                    f"펀진_사업보고서_2026.03_{statement_name}.xlsx",
                )
                workbook = load_workbook(path)
                self.assertEqual(workbook.sheetnames, [statement_name])


class MultiYearMergeTests(unittest.TestCase):
    def test_merges_by_conservative_item_identity_and_uses_latest_overlap(self):
        old_frame = pd.DataFrame(
            [
                ["Ⅰ. 유동자산", 100, 90],
                ["1. 현금 (주석 3)", 30, 20],
                ["2. 매출채권", 70, 70],
            ],
            columns=["항목", "2024-12-31", "2023-12-31"],
        )
        new_frame = pd.DataFrame(
            [
                ["유동자산", 130, 110],
                ["1. 현금및현금성자산 (주석 4)", 40, 35],
                ["2. 계약자산", 10, 5],
                ["3. 매출채권", 80, 75],
            ],
            columns=["항목", "2025-12-31", "2024-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        self.assertEqual(
            merged.columns.to_list(),
            ["항목", "2025-12-31", "2024-12-31", "2023-12-31"],
        )
        receivables = merged[merged["항목"] == "3. 매출채권"]
        self.assertEqual(len(receivables), 1)
        self.assertEqual(receivables.iloc[0]["2024-12-31"], 75)
        self.assertEqual(receivables.iloc[0]["2023-12-31"], 70)

        self.assertIn("1. 현금 (주석 3)", merged["항목"].to_list())
        self.assertIn("1. 현금및현금성자산 (주석 4)", merged["항목"].to_list())
        old_cash = merged[merged["항목"] == "1. 현금 (주석 3)"].iloc[0]
        new_cash = merged[
            merged["항목"] == "1. 현금및현금성자산 (주석 4)"
        ].iloc[0]
        self.assertEqual(old_cash["2023-12-31"], 20)
        self.assertEqual(old_cash["2024-12-31"], "")
        self.assertEqual(new_cash["2024-12-31"], 35)
        self.assertEqual(new_cash["2023-12-31"], "")

    def test_note_reference_change_does_not_create_duplicate_row(self):
        old_frame = pd.DataFrame(
            [["1. 현금및현금성자산 (주석 3과 6)", 30]],
            columns=["항목", "2024-12-31"],
        )
        new_frame = pd.DataFrame(
            [["현금및현금성자산 (주석 4, 7)", 40]],
            columns=["항목", "2025-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        self.assertEqual(len(merged), 1)
        self.assertEqual(merged.iloc[0]["2025-12-31"], 40)
        self.assertEqual(merged.iloc[0]["2024-12-31"], 30)

    def test_merges_known_label_aliases_within_the_same_section(self):
        old_frame = pd.DataFrame(
            [
                ["Ⅰ.매출액", 100],
                ["제품매출액", 100],
                ["Ⅱ.판매비와관리비", 10],
                ["세금과공과", 10],
                ["Ⅴ.영업이익", 90],
                ["Ⅹ.당기순이익", 90],
            ],
            columns=["항목", "2024-12-31"],
        )
        new_frame = pd.DataFrame(
            [
                ["Ⅰ.매출액", 120],
                ["1. 제품매출", 120],
                ["Ⅱ.판매비와관리비", 12],
                ["9. 세금과공과금", 12],
                ["Ⅴ.영업이익(손실)", 108],
                ["Ⅹ.당기순이익(손실)", 108],
            ],
            columns=["항목", "2025-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        for label in (
            "1. 제품매출",
            "9. 세금과공과금",
            "Ⅴ.영업이익(손실)",
            "Ⅹ.당기순이익(손실)",
        ):
            row = merged[merged["항목"] == label]
            self.assertEqual(len(row), 1)
            self.assertNotEqual(row.iloc[0]["2024-12-31"], "")
            self.assertNotEqual(row.iloc[0]["2025-12-31"], "")

    def test_merges_government_subsidy_only_under_the_same_parent(self):
        old_frame = pd.DataFrame(
            [
                ["자 산", "",],
                ["Ⅱ.비유동자산", 270],
                ["(2) 유형자산", 270],
                ["1. 토지", 100],
                ["국고보조금", -10],
                ["2. 건물", 200],
                ["국고보조금", -20],
            ],
            columns=["항목", "2024-12-31"],
        )
        new_frame = pd.DataFrame(
            [
                ["자 산", ""],
                ["Ⅱ.비유동자산", 268],
                ["(2) 유형자산", 268],
                ["1. 토지", 100],
                ["정부보조금", -11],
                ["2. 건물", 200],
                ["정부보조금", -21],
            ],
            columns=["항목", "2025-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        subsidies = merged[merged["항목"] == "정부보조금"].reset_index(drop=True)
        self.assertEqual(len(subsidies), 2)
        self.assertEqual(subsidies.iloc[0]["2025-12-31"], -11)
        self.assertEqual(subsidies.iloc[0]["2024-12-31"], -10)
        self.assertEqual(subsidies.iloc[1]["2025-12-31"], -21)
        self.assertEqual(subsidies.iloc[1]["2024-12-31"], -20)

    def test_reconciles_a_same_item_that_moved_within_its_section(self):
        old_frame = pd.DataFrame(
            [
                ["Ⅰ.유동자산", 15],
                ["미수수익", 10],
                ["선급비용", 5],
            ],
            columns=["항목", "2024-12-31"],
        )
        new_frame = pd.DataFrame(
            [
                ["Ⅰ.유동자산", 26],
                ["8. 선급비용", 6],
                ["6. 미수수익(주석14)", 20],
            ],
            columns=["항목", "2025-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        interest = merged[merged["항목"] == "6. 미수수익(주석14)"]
        prepaid = merged[merged["항목"] == "8. 선급비용"]
        self.assertEqual(len(interest), 1)
        self.assertEqual(interest.iloc[0]["2024-12-31"], 10)
        self.assertEqual(interest.iloc[0]["2025-12-31"], 20)
        self.assertEqual(len(prepaid), 1)
        self.assertEqual(prepaid.iloc[0]["2024-12-31"], 5)
        self.assertEqual(prepaid.iloc[0]["2025-12-31"], 6)

    def test_merges_clear_pdf_line_break_and_cash_flow_label_variants(self):
        old_frame = pd.DataFrame(
            [
                ["Ⅰ.영업활동으로 인한 현금흐름", 100],
                ["2.현금의 유출이 없는 비용등의 가산", 20],
                ["4.영업활동으로 인한 자산부채의 변동", 30],
                ["가. 매출채권의 감소(증가)", 50],
            ],
            columns=["항목", "2024-12-31"],
        )
        new_frame = pd.DataFrame(
            [
                ["Ⅰ.영업활동으로 인한 현금흐름", 120],
                ["2. 현금의 유출이 없는 비용등의 가", 25],
                ["4. 영업활동으로 인한 자산부채의 변", 35],
                ["가. 매출채권의 증가", 60],
            ],
            columns=["항목", "2025-12-31"],
        )

        merged = merge_statement_frames(
            [
                (FilingMetadata("테스트", "사업보고서", "2025.03"), old_frame),
                (FilingMetadata("테스트", "사업보고서", "2026.03"), new_frame),
            ]
        )

        for label in (
            "2.현금의 유출이 없는 비용등의 가산",
            "4.영업활동으로 인한 자산부채의 변동",
            "가. 매출채권의 증가",
        ):
            row = merged[merged["항목"] == label]
            self.assertEqual(len(row), 1)
            self.assertNotEqual(row.iloc[0]["2024-12-31"], "")
            self.assertNotEqual(row.iloc[0]["2025-12-31"], "")

    def test_old_audit_pdf_uses_current_year_and_removes_document_headers(self):
        frame = pd.DataFrame(
            [
                ["재 무 상 태 표", "", ""],
                ["제 기 2022년 12월 31일 현재", 16, ""],
                ["제 기 2021년 12월 31일 현재", 15, ""],
                ["회사명 : 주식회사 테스트 (단위 : 원)", "", ""],
                ["제 (전) 기", "", 15],
                ["과 목 제 (당) 기", 16, ""],
                ["자 산", "", ""],
                ["Ⅰ.유동자산", 100, 90],
                ["", 200, ""],
                ["자 산 총 계", "", 180],
                ["#NAME?", "", ""],
                ["별첨 재무제표에 대한 주석 참조", "", ""],
            ],
            columns=["과목", "당기", "전기"],
        )

        merged = merge_statement_frames(
            [(FilingMetadata("테스트", "감사보고서", "2023.04"), frame)],
            primary_period_only=True,
        )

        self.assertEqual(merged.columns.to_list(), ["항목", "2022년"])
        self.assertEqual(
            merged["항목"].to_list(),
            ["자 산", "Ⅰ.유동자산", "자 산 총 계"],
        )
        total = merged[merged["항목"] == "자 산 총 계"].iloc[0]
        self.assertEqual(total["2022년"], 200)

    def test_quarterly_merge_keeps_only_the_current_standalone_quarter(self):
        frame = pd.DataFrame(
            [["매출액", 30, 90, 25, 70]],
            columns=[
                "항목",
                "2025-07-01 ~ 2025-09-30",
                "2025-01-01 ~ 2025-09-30",
                "2024-07-01 ~ 2024-09-30",
                "2024-01-01 ~ 2024-09-30",
            ],
        )

        merged = merge_statement_frames(
            [(FilingMetadata("테스트", "분기보고서", "2025.11"), frame)],
            primary_period_only=True,
        )

        self.assertEqual(merged.columns.to_list(), ["항목", "2025년 3분기"])
        self.assertEqual(merged.iloc[0]["2025년 3분기"], 30)

    def test_run_many_creates_one_multi_year_workbook(self):
        def filing(receipt, year_month, current_year, prior_year):
            statements = {
                name: pd.DataFrame(
                    [["자산", current_year, prior_year]],
                    columns=[
                        "항목",
                        f"{current_year}-12-31",
                        f"{prior_year}-12-31",
                    ],
                )
                for name in STATEMENT_NAMES
            }
            return FilingExtraction(
                metadata=FilingMetadata("테스트 주식회사", "사업보고서", year_month),
                statements=statements,
                selected_scope="consolidated",
                receipt_number=receipt,
                source="XBRL",
            )

        extractions = [
            filing("20240319000709", "2024.03", 2023, 2022),
            filing("20250319000710", "2025.03", 2024, 2023),
        ]
        dart = SimpleNamespace(set_api_key=lambda **kwargs: None)

        with tempfile.TemporaryDirectory() as temp_dir:
            with patch(
                "extract_financials._extract_filing",
                side_effect=extractions,
            ):
                paths, returned_filings = run_many(
                    "20240319000709,\n20250319000710",
                    output_dir=Path(temp_dir),
                    environ={"DART_API_KEY": "secret"},
                    dart_module=dart,
                )

            self.assertEqual(len(returned_filings), 2)
            self.assertEqual(
                paths[0].name,
                "테스트 주식회사_사업보고서_2023-2024.xlsx",
            )
            workbook = load_workbook(paths[0])
            self.assertEqual(workbook.sheetnames, list(STATEMENT_NAMES))
            self.assertEqual(
                [cell.value for cell in workbook["재무상태표"][1]],
                ["항목", "2024년", "2023년"],
            )

    def test_run_many_rejects_different_companies(self):
        frame = pd.DataFrame([["자산", 1]], columns=["항목", "2025-12-31"])
        statements = {name: frame for name in STATEMENT_NAMES}
        extractions = [
            FilingExtraction(
                FilingMetadata("회사 A", "사업보고서", "2025.03"),
                statements,
                "consolidated",
                "20240319000709",
                "XBRL",
            ),
            FilingExtraction(
                FilingMetadata("회사 B", "사업보고서", "2026.03"),
                statements,
                "consolidated",
                "20250319000710",
                "XBRL",
            ),
        ]
        dart = SimpleNamespace(set_api_key=lambda **kwargs: None)

        with (
            patch("extract_financials._extract_filing", side_effect=extractions),
            self.assertRaisesRegex(FinancialExtractionError, "같은 회사"),
        ):
            run_many(
                "20240319000709,20250319000710",
                environ={"DART_API_KEY": "secret"},
                dart_module=dart,
            )

    def test_run_many_rejects_mixed_scopes(self):
        frame = pd.DataFrame([["자산", 1]], columns=["항목", "2025-12-31"])
        statements = {name: frame for name in STATEMENT_NAMES}
        extractions = [
            FilingExtraction(
                FilingMetadata("같은 회사", "사업보고서", "2025.03"),
                statements,
                "consolidated",
                "20240319000709",
                "XBRL",
            ),
            FilingExtraction(
                FilingMetadata("같은 회사", "사업보고서", "2026.03"),
                statements,
                "separate",
                "20250319000710",
                "PDF",
            ),
        ]
        dart = SimpleNamespace(set_api_key=lambda **kwargs: None)

        with (
            patch("extract_financials._extract_filing", side_effect=extractions),
            self.assertRaisesRegex(FinancialExtractionError, "섞을 수 없습니다"),
        ):
            run_many(
                "20240319000709,20250319000710",
                environ={"DART_API_KEY": "secret"},
                dart_module=dart,
            )


class PipelineFallbackTests(unittest.TestCase):
    def test_run_uses_pdf_only_when_xbrl_is_missing(self):
        finance = SimpleNamespace(
            download_xbrl=lambda **kwargs: (_ for _ in ()).throw(
                FileNotFoundError("XBRL File Not Found")
            )
        )
        dart = SimpleNamespace(
            api=SimpleNamespace(finance=finance),
            set_api_key=lambda **kwargs: None,
        )
        statements = {
            name: pd.DataFrame([["현금", 1_000]], columns=["과목", "당기"])
            for name in STATEMENT_NAMES
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            with (
                patch(
                    "extract_financials.resolve_document_number",
                    return_value="11134296",
                ),
                patch(
                    "extract_financials.download_filing_pdf",
                    return_value=Path(temp_dir) / "filing.pdf",
                ),
                patch(
                    "extract_financials.extract_pdf_statements",
                    return_value=(statements, "consolidated"),
                ),
                patch(
                    "extract_financials.resolve_filing_metadata",
                    return_value=FilingMetadata(
                        "넥슨",
                        "감사보고서",
                        "2026.05",
                    ),
                ),
            ):
                paths, selected_scope, receipt_number, source = run(
                    "20260317801285",
                    output_dir=Path(temp_dir) / "output",
                    environ={"DART_API_KEY": "secret"},
                    dart_module=dart,
                )

        self.assertEqual(len(paths), 1)
        self.assertEqual(selected_scope, "consolidated")
        self.assertEqual(receipt_number, "20260317801285")
        self.assertEqual(source, "PDF")
        self.assertEqual(paths[0].name, "넥슨_감사보고서_2026.05.xlsx")


if __name__ == "__main__":
    unittest.main()
